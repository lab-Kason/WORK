#Stable/dun move
import streamlit as st
import csv
import os
import sys
import tempfile
from PyPDF2 import PdfReader  # For reading PDF files
from docx import Document  # For reading .docx files
import xlrd  # For reading .xls files
from openpyxl import load_workbook  # For reading .xlsx files

# Increase recursion limit
sys.setrecursionlimit(5000)

# File extraction functions
def extract_text_from_pdf(pdf_path):
    try:
        # Attempt extraction with PyPDF2
        reader = PdfReader(pdf_path)
        text = ""
        for page in reader.pages:
            text += page.extract_text() + "\n"
        return text
    except Exception as e:
        st.error(f"Could not read PDF file {pdf_path}: {e}")
        return ""

def extract_text_from_txt(txt_path):
    try:
        with open(txt_path, 'r', encoding='utf-8') as file:
            return file.read()
    except Exception as e:
        st.error(f"Could not read TXT file {txt_path}: {e}")
        return ""

def extract_text_from_docx(docx_path):
    try:
        doc = Document(docx_path)
        return "\n".join([paragraph.text for paragraph in doc.paragraphs])
    except Exception as e:
        st.error(f"Could not read DOCX file {docx_path}: {e}")
        return ""

def extract_text_from_xls(xls_path):
    try:
        workbook = xlrd.open_workbook(xls_path)
        data = []
        for sheet in workbook.sheets():
            for row_idx in range(sheet.nrows):
                row = sheet.row_values(row_idx)
                data.append(row)
        return data
    except Exception as e:
        st.error(f"Could not read XLS file {xls_path}: {e}")
        return []

def extract_text_from_xlsx(xlsx_path):
    try:
        workbook = load_workbook(xlsx_path, data_only=True)
        text = ""
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for row in worksheet.iter_rows(values_only=True):
                text += " ".join([str(cell) for cell in row if cell is not None]) + "\n"
        return text
    except Exception as e:
        st.error(f"Could not read XLSX file {xlsx_path}: {e}")
        return ""

def extract_text(file_path):
    if file_path.lower().endswith(".pdf"):
        return extract_text_from_pdf(file_path)
    elif file_path.lower().endswith(".txt"):
        return extract_text_from_txt(file_path)
    elif file_path.lower().endswith(".docx"):
        return extract_text_from_docx(file_path)
    elif file_path.lower().endswith(".xls"):
        return extract_text_from_xls(file_path)
    elif file_path.lower().endswith(".xlsx"):
        return extract_text_from_xlsx(file_path)
    else:
        st.error(f"Unsupported file type: {file_path}")
        return ""

# Data extraction function
def extract_data_from_pdf(text, keywords, behaviors):
    extracted_data = {}
    meaningless_words = {"Attachments", "Page", "Document", "File"}  # Define meaningless words

    if isinstance(text, list):  # Handle .xls data (list of rows)
        for column, keyword in keywords.items():
            behavior = behaviors.get(column, "right")
            values = []  # Collect all matches
            for row_idx, row in enumerate(text):
                if keyword in row:
                    keyword_idx = row.index(keyword)
                    if behavior == "right":
                        value = row[keyword_idx + 1] if keyword_idx + 1 < len(row) else "N/A"
                    elif behavior == "left":
                        value = row[keyword_idx - 1] if keyword_idx - 1 >= 0 else "N/A"
                    elif behavior == "below":
                        for next_row_idx in range(row_idx + 1, len(text)):
                            next_row = text[next_row_idx]
                            if next_row[keyword_idx]:
                                value = next_row[keyword_idx]
                                if value not in values:  # Avoid duplicates
                                    values.append(value)
                        continue
                    elif behavior == "above":
                        for prev_row_idx in range(row_idx - 1, -1, -1):
                            prev_row = text[prev_row_idx]
                            if prev_row[keyword_idx]:
                                value = prev_row[keyword_idx]
                                if value not in values:  # Avoid duplicates
                                    values.append(value)
                        continue
                    elif behavior == "keyword":
                        value = keyword
                    if value not in values:  # Avoid duplicates
                        values.append(value)
            extracted_data[column] = values if values else ["N/A"]
    else:  # Handle text data (e.g., PDF, TXT, DOCX)
        lines = text.split("\n")
        for column, keyword in keywords.items():
            behavior = behaviors.get(column, "right")
            values = []  # Collect all matches
            for i, line in enumerate(lines):
                if keyword in line:
                    if behavior == "right":
                        start_index = line.find(keyword) + len(keyword)
                        remaining_text = line[start_index:].strip()
                        # Extract meaningful text until encountering meaningless words
                        meaningful_text = []
                        for word in remaining_text.split():
                            if word in meaningless_words:
                                break  # Stop if encountering meaningless words
                            meaningful_text.append(word)
                        value = " ".join(meaningful_text) if meaningful_text else "N/A"
                    elif behavior == "left":
                        start_index = line.find(keyword)
                        preceding_text = line[:start_index].strip()
                        meaningful_text = []
                        for word in reversed(preceding_text.split()):  # Reverse to check left
                            if word in meaningless_words:
                                break  # Stop if encountering meaningless words
                            meaningful_text.insert(0, word)
                        value = " ".join(meaningful_text) if meaningful_text else "N/A"
                    elif behavior == "below":
                        for next_line_idx in range(i + 1, len(lines)):
                            next_line = lines[next_line_idx].strip()
                            if next_line:
                                meaningful_text = []
                                for word in next_line.split():
                                    if word in meaningless_words:
                                        break  # Stop if encountering meaningless words
                                    meaningful_text.append(word)
                                value = " ".join(meaningful_text) if meaningful_text else "N/A"
                                if value not in values:  # Avoid duplicates
                                    values.append(value)
                        continue
                    elif behavior == "above":
                        for prev_line_idx in range(i - 1, -1, -1):
                            prev_line = lines[prev_line_idx].strip()
                            if prev_line:
                                meaningful_text = []
                                for word in prev_line.split():
                                    if word in meaningless_words:
                                        break  # Stop if encountering meaningless words
                                    meaningful_text.append(word)
                                value = " ".join(meaningful_text) if meaningful_text else "N/A"
                                if value not in values:  # Avoid duplicates
                                    values.append(value)
                        continue
                    elif behavior == "keyword":
                        value = keyword
                    if value not in values:  # Avoid duplicates
                        values.append(value)
            # Remove meaningless words from the extracted result
            extracted_data[column] = [
                val for val in values if val not in meaningless_words
            ] if values else ["N/A"]
    return extracted_data

# Streamlit app
def main():
    st.title("File Data Extraction and CSV Generator")
    
    # File upload
    uploaded_files = st.file_uploader("Upload files", accept_multiple_files=True)
    
    # Column titles and keywords
    column_titles = st.text_input("Enter column titles (comma-separated)", "Item,Description,Qty,Amount")
    column_titles = [title.strip() for title in column_titles.split(",")]
    
    keywords = {}
    for column in column_titles:
        keywords[column] = st.text_input(f"Enter keyword for column '{column}'", column)
    
    # Extraction behaviors
    extraction_behaviors = {}
    for column in column_titles:
        extraction_behaviors[column] = st.selectbox(
            f"Select extraction behavior for column '{column}'",
            ["right", "left", "below", "above", "keyword"],  # Added "keyword" option
            index=2
        )
    
    # Process files and generate CSV
    if st.button("Generate CSV"):
        if uploaded_files:
            rows = []
            for uploaded_file in uploaded_files:
                # Save the uploaded file to a temporary directory
                file_extension = os.path.splitext(uploaded_file.name)[1]
                with tempfile.NamedTemporaryFile(delete=False, suffix=file_extension) as temp_file:
                    temp_file.write(uploaded_file.read())
                    temp_file_path = temp_file.name
                
                # Extract text from the temporary file
                text = extract_text(temp_file_path)
                if not text:
                    st.error(f"Failed to extract text from file: {uploaded_file.name}")
                    continue
                
                # Extract data from the text
                extracted_data = extract_data_from_pdf(text, keywords, extraction_behaviors)
                row = [", ".join(extracted_data.get(column, ["N/A"])) for column in column_titles]
                rows.append(row)
            
            # Save to CSV
            csv_file_path = os.path.join(os.getcwd(), "output.csv")
            with open(csv_file_path, mode="w", newline="") as csv_file:
                writer = csv.writer(csv_file)
                writer.writerow(column_titles)
                writer.writerows(rows)
            
            # Provide download button for the CSV file
            with open(csv_file_path, "rb") as f:
                st.download_button(
                    label="Download CSV",
                    data=f.read(),
                    file_name="output.csv",
                    mime="text/csv"
                )
        else:
            st.error("No files uploaded!")

if __name__ == "__main__":
    main()
