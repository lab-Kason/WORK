import csv
import os
from AppKit import NSOpenPanel
import pdfplumber  # For reading PDF files
import re
from docx import Document  # For reading .docx files
import xlrd  # For reading .xls files
from openpyxl import load_workbook  # For reading .xlsx files

def select_files_or_folders():
    """
    Opens a Finder dialog to let the user select multiple files or folders.

    Returns:
        list: A list of selected file or folder paths.
    """
    panel = NSOpenPanel.openPanel()
    panel.setAllowsMultipleSelection_(True)  # Allow multiple selections
    panel.setCanChooseDirectories_(True)    # Allow selecting directories
    panel.setCanChooseFiles_(True)          # Allow selecting files
    panel.setTitle_("Select Files or Folders")  # Set the title of the dialog

    if panel.runModal() == 1:  # If the user clicks "Open"
        return [str(url.path()) for url in panel.URLs()]  # Get the selected paths
    return []  # Return an empty list if the user cancels

def extract_text_from_pdf(pdf_path):
    """
    Extracts text from a PDF file using pdfplumber.

    Args:
        pdf_path (str): The path to the PDF file.

    Returns:
        str: The extracted text from the PDF.
    """
    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = ""
            for page in pdf.pages:
                text += page.extract_text()  # Extract text from each page
            return text
    except Exception as e:
        print(f"Could not read PDF file {pdf_path}: {e}")
        return ""

def extract_text_from_txt(txt_path):
    """
    Extracts text from a .txt file.

    Args:
        txt_path (str): The path to the .txt file.

    Returns:
        str: The extracted text from the .txt file.
    """
    try:
        with open(txt_path, 'r', encoding='utf-8') as file:
            return file.read()
    except Exception as e:
        print(f"Could not read TXT file {txt_path}: {e}")
        return ""

def extract_text_from_docx(docx_path):
    """
    Extracts text from a .docx file.

    Args:
        docx_path (str): The path to the .docx file.

    Returns:
        str: The extracted text from the .docx file.
    """
    try:
        doc = Document(docx_path)
        return "\n".join([paragraph.text for paragraph in doc.paragraphs])
    except Exception as e:
        print(f"Could not read DOCX file {docx_path}: {e}")
        return ""

def extract_text_from_xls(xls_path):
    """
    Extracts text from an .xls file using xlrd.

    Args:
        xls_path (str): The path to the .xls file.

    Returns:
        str: The extracted text from the .xls file.
    """
    try:
        workbook = xlrd.open_workbook(xls_path)
        text = ""
        for sheet in workbook.sheets():
            for row_idx in range(sheet.nrows):
                row = sheet.row_values(row_idx)
                text += " ".join([str(cell) for cell in row if cell]) + "\n"
        return text
    except Exception as e:
        print(f"Could not read XLS file {xls_path}: {e}")
        return ""

def extract_text_from_xlsx(xlsx_path):
    """
    Extracts text from an .xlsx file using openpyxl.

    Args:
        xlsx_path (str): The path to the .xlsx file.

    Returns:
        str: The extracted text from the .xlsx file.
    """
    try:
        workbook = load_workbook(xlsx_path, data_only=True)
        text = ""
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for row in worksheet.iter_rows(values_only=True):
                text += " ".join([str(cell) for cell in row if cell is not None]) + "\n"
        return text
    except Exception as e:
        print(f"Could not read XLSX file {xlsx_path}: {e}")
        return ""

def extract_text(file_path):
    """
    Extracts text from a file based on its type.

    Args:
        file_path (str): The path to the file.

    Returns:
        str: The extracted text from the file.
    """
    if file_path.lower().endswith(".pdf"):
        return extract_text_from_pdf(file_path)
    elif file_path.lower().endswith(".txt"):
        return extract_text_from_txt(file_path)
    elif file_path.lower().endswith(".docx"):
        return extract_text_from_docx(file_path)
    elif file_path.lower().endswith(".xls"):
        return extract_text_from_xls(file_path)
    elif file_path.lower().endswith(".xlsx"):
        return extract_text_from_xlsx(file_path)
    else:
        print(f"Unsupported file type: {file_path}")
        return ""

def prompt_for_columns_and_references():
    """
    Prompts the user to input column titles, keywords, and select files/folders for each column.

    Returns:
        list: A list of column titles.
        dict: A dictionary mapping column titles to their corresponding keywords.
        dict: A dictionary mapping column titles to their selected files/folders.
        dict: A dictionary mapping column titles to their extraction source (title or content).
    """
    print("Enter the column titles separated by commas (e.g., Name, Age, Address):")
    column_titles = input().strip().split(",")
    column_titles = [title.strip() for title in column_titles]  # Clean up whitespace

    keywords = {}
    references = {}
    extraction_sources = {}

    for column in column_titles:
        # Prompt for the keyword for the column
        keyword = input(f"Enter the keyword to search for the column '{column}': ").strip()
        keywords[column] = keyword

        # Ask the user if the keyword should be extracted from the file title or content
        while True:
            source = input(f"Should the keyword for '{column}' be extracted from the file title or content? (Enter 'title' or 'content'): ").strip().lower()
            if source in ["title", "content"]:
                extraction_sources[column] = source
                break
            else:
                print("Invalid input. Please enter 'title' or 'content'.")

        # Prompt the user to select files or folders for the column
        print(f"Select files or folders for the column '{column}':")
        selected_paths = select_files_or_folders()
        if selected_paths:
            references[column] = selected_paths
        else:
            print(f"No files or folders selected for column '{column}'.")
            references[column] = []

    return column_titles, keywords, references, extraction_sources

def extract_data_from_pdf(text, keywords):
    """
    Extracts data from the text based on the provided keywords.

    Args:
        text (str): The extracted text from the file.
        keywords (dict): A dictionary mapping column titles to keywords.

    Returns:
        dict: A dictionary mapping column titles to extracted values.
    """
    # Normalize the text by removing extra spaces and newlines
    normalized_text = " ".join(text.split())
    print(f"Normalized text:\n{normalized_text}")  # Debug: Print the normalized text

    extracted_data = {}
    for column, keyword in keywords.items():
        if keyword in normalized_text:
            # Extract the value after the keyword
            start_index = normalized_text.find(keyword) + len(keyword)
            remaining_text = normalized_text[start_index:].strip()
            
            # Handle cases where the value follows immediately after the keyword
            value = remaining_text.split()[0] if remaining_text else "N/A"
            
            # Validate the extracted value
            if value == ":":
                # If the extracted value is just a colon, try extracting the next part
                remaining_text = remaining_text[1:].strip()  # Skip the colon
                value = remaining_text.split()[0] if remaining_text else "N/A"
            
            extracted_data[column] = value
        else:
            # If the keyword is not found, set the value to "N/A"
            extracted_data[column] = "N/A"
    return extracted_data

def process_columns_and_generate_csv(column_titles, keywords, references, extraction_sources, csv_file_path):
    """
    Processes the selected files for each column and generates a CSV file with extracted data.

    Args:
        column_titles (list): A list of column titles.
        keywords (dict): A dictionary mapping column titles to keywords.
        references (dict): A dictionary mapping column titles to their selected files/folders.
        extraction_sources (dict): A dictionary mapping column titles to their extraction source (title or content).
        csv_file_path (str): The path to the output CSV file.
    """
    rows = []

    # Process each column independently
    for column in column_titles:
        column_keyword = keywords[column]
        column_references = references[column]
        extraction_source = extraction_sources[column]

        for path in column_references:
            if os.path.isfile(path):
                if extraction_source == "title":
                    # Directly use the file title as the extracted result
                    file_name = os.path.basename(path)
                    value = file_name  # Use the full file name as the value
                else:
                    # Extract from the file content
                    text = extract_text(path)
                    extracted_data = extract_data_from_pdf(text, {column: column_keyword})
                    value = extracted_data[column] if column in extracted_data else "N/A"

                row = [value if col == column else "N/A" for col in column_titles]
                rows.append(row)
            elif os.path.isdir(path):
                items = os.listdir(path)
                for item in items:
                    item_path = os.path.join(path, item)
                    if os.path.isfile(item_path):
                        if extraction_source == "title":
                            # Directly use the file title as the extracted result
                            file_name = os.path.basename(item_path)
                            value = file_name  # Use the full file name as the value
                        else:
                            # Extract from the file content
                            text = extract_text(item_path)
                            extracted_data = extract_data_from_pdf(text, {column: column_keyword})
                            value = extracted_data[column] if column in extracted_data else "N/A"

                        row = [value if col == column else "N/A" for col in column_titles]
                        rows.append(row)

    # Write the rows to the CSV file
    try:
        with open(csv_file_path, mode="w", newline="") as csv_file:
            writer = csv.writer(csv_file)
            writer.writerow(column_titles)  # Write the header row
            writer.writerows(rows)  # Write the data rows
        print(f"CSV file created successfully at {csv_file_path}")
    except Exception as e:
        print(f"An error occurred while writing to the CSV file: {e}")

def generate_csv():
    # Get the desktop path
    desktop_path = os.path.expanduser("~/Desktop")
    
    # Prompt the user for the CSV file name
    csv_file_name = input("Enter the name of the CSV file (e.g., output.csv): ").strip()
    
    # Ensure the file name ends with .csv
    if not csv_file_name.lower().endswith(".csv"):
        csv_file_name += ".csv"
    
    # Combine the desktop path with the file name
    csv_file_path = os.path.join(desktop_path, csv_file_name)
    
    # Prompt the user for column titles, keywords, and references
    column_titles, keywords, references, extraction_sources = prompt_for_columns_and_references()
    
    # Process the columns and generate the CSV
    process_columns_and_generate_csv(column_titles, keywords, references, extraction_sources, csv_file_path)

# Run the program
if __name__ == "__main__":
    generate_csv()
